from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any


_A1_RANGE_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+$")
_SHEET_BOUNDS_FALLBACK_MAX = 5000


def _is_number(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return False
        s = s.replace(",", "")
        try:
            float(s)
            return True
        except Exception:
            return False
    return False


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Keep readable numeric strings for LLM + markdown.
        if abs(value) >= 1e12 or (abs(value) > 0 and abs(value) < 1e-6):
            return f"{value:.6g}"
        # Strip trailing zeros.
        s = f"{value:.10f}".rstrip("0").rstrip(".")
        return s if s else "0"
    return str(value)


def table_to_csv(rows: list[list[str]]) -> str:
    buf = StringIO()
    w = csv.writer(buf)
    for r in rows:
        w.writerow([c if c is not None else "" for c in (r or [])])
    return buf.getvalue()


def table_to_markdown(rows: list[list[str]], *, max_rows: int = 40, max_cols: int = 16) -> str:
    if not rows:
        return ""
    trimmed = [list(r[:max_cols]) for r in rows[:max_rows]]
    cols = max((len(r) for r in trimmed), default=0)
    if cols <= 0:
        return ""
    normalized = [r + [""] * (cols - len(r)) for r in trimmed]

    header = normalized[0]
    body = normalized[1:] if len(normalized) > 1 else []
    lines: list[str] = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * cols) + " |")
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    suffix = ""
    if len(rows) > max_rows:
        suffix += f"\n… ({len(rows) - max_rows} more rows)"
    if any(len(r) > max_cols for r in rows[:max_rows]):
        suffix += f"\n… (some rows truncated to {max_cols} cols)"
    return "\n".join(lines).strip() + suffix


def _trim_empty(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return []

    # Trim empty rows at bottom.
    end = len(rows)
    while end > 0 and all((c or "").strip() == "" for c in rows[end - 1]):
        end -= 1
    rows = rows[:end]
    if not rows:
        return []

    # Trim empty columns at right.
    max_cols = max((len(r) for r in rows), default=0)
    if max_cols <= 0:
        return []
    last_non_empty = -1
    for c in range(max_cols - 1, -1, -1):
        if any((c < len(r) and (r[c] or "").strip() != "") for r in rows):
            last_non_empty = c
            break
    if last_non_empty == -1:
        return []
    return [r[: last_non_empty + 1] for r in rows]


def validate_a1_range(a1_range: str) -> bool:
    s = (a1_range or "").strip().upper().replace(" ", "")
    return bool(_A1_RANGE_RE.match(s))


@dataclass(frozen=True)
class ExcelBlockCandidate:
    sheet: str
    a1_range: str
    numeric_cells: int
    total_cells: int
    preview_rows: list[list[str]]


def load_workbook_bytes(xlsx_bytes: bytes):
    from openpyxl import load_workbook

    return load_workbook(BytesIO(xlsx_bytes), data_only=True)


def extract_a1_range(ws, a1_range: str, *, max_rows: int = 200, max_cols: int = 40) -> list[list[str]]:
    from openpyxl.utils.cell import range_boundaries

    a1 = (a1_range or "").strip().upper().replace(" ", "")
    min_col, min_row, max_col, max_row = range_boundaries(a1)
    max_row = min(max_row, min_row + max_rows - 1)
    max_col = min(max_col, min_col + max_cols - 1)

    rows: list[list[str]] = []
    for r in range(min_row, max_row + 1):
        out_row: list[str] = []
        for c in range(min_col, max_col + 1):
            out_row.append(_to_str(ws.cell(row=r, column=c).value))
        rows.append(out_row)
    return _trim_empty(rows)


def sheet_preview(ws, *, rows: int = 18, cols: int = 10) -> list[list[str]]:
    # A1 origin preview: good for LLM selection + debugging.
    preview: list[list[str]] = []
    header = [""] + [f"{i}" for i in range(1, cols + 1)]
    preview.append(header)
    for r in range(1, rows + 1):
        row = [str(r)]
        for c in range(1, cols + 1):
            row.append(_to_str(ws.cell(row=r, column=c).value))
        preview.append(row)
    return preview


def _numeric_cells(ws) -> list[tuple[int, int]]:
    """
    Return (row, col) for numeric-valued cells, using ws._cells when available.
    This avoids scanning massive empty regions caused by formatting.
    """
    out: list[tuple[int, int]] = []
    cells = getattr(ws, "_cells", None)
    if isinstance(cells, dict) and cells:
        for (row, col), cell in cells.items():
            try:
                if _is_number(cell.value):
                    out.append((int(row), int(col)))
            except Exception:
                continue
        return out

    # Fallback: bounded scan from top-left.
    max_row = min(int(getattr(ws, "max_row", 1) or 1), _SHEET_BOUNDS_FALLBACK_MAX)
    max_col = min(int(getattr(ws, "max_column", 1) or 1), 80)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            try:
                if _is_number(ws.cell(row=r, column=c).value):
                    out.append((r, c))
            except Exception:
                continue
    return out


def _dedupe_candidates(cands: list[ExcelBlockCandidate]) -> list[ExcelBlockCandidate]:
    seen: set[tuple[str, str]] = set()
    out: list[ExcelBlockCandidate] = []
    for c in cands:
        key = (c.sheet, c.a1_range)
        if key in seen:
            continue
        seen.add(key)
        out.append(c)
    return out


def _find_numeric_blocks_in_window(
    ws,
    *,
    start_row: int,
    start_col: int,
    height: int,
    width: int,
    max_blocks: int,
) -> list[ExcelBlockCandidate]:
    """
    Heuristic numeric-block detection for MVP:
    - Find contiguous-ish numeric rectangles (often measurement tables).
    - Include a header row if present (string row just above numeric block).
    """
    from openpyxl.utils.cell import get_column_letter

    max_r = max(1, height)
    max_c = max(1, width)
    grid: list[list[Any]] = []
    for r in range(start_row, start_row + max_r):
        row: list[Any] = []
        for c in range(start_col, start_col + max_c):
            row.append(ws.cell(row=r, column=c).value)
        grid.append(row)

    numeric = [[_is_number(v) for v in row] for row in grid]

    candidates: list[ExcelBlockCandidate] = []
    visited: set[tuple[int, int]] = set()

    for r in range(max_r):
        for c in range(max_c):
            if not numeric[r][c] or (r, c) in visited:
                continue

            # Expand width on this row.
            width = 0
            while c + width < max_c and numeric[r][c + width]:
                width += 1
            if width < 2:
                continue

            # Expand height while rows keep enough numeric coverage.
            height = 1
            while r + height < max_r:
                row_cov = sum(1 for k in range(width) if numeric[r + height][c + k])
                if row_cov / float(width) < 0.6:
                    break
                height += 1

            if height < 2:
                continue

            # Mark visited numeric cells in the rectangle (best-effort).
            for rr in range(r, r + height):
                for cc in range(c, c + width):
                    if numeric[rr][cc]:
                        visited.add((rr, cc))

            min_row = start_row + r
            min_col = start_col + c
            max_row = start_row + (r + height - 1)
            max_col = start_col + (c + width - 1)

            # Optional header row above.
            # Some sheets place the header several rows above the densest numeric block,
            # and sometimes the left-most header is slightly left of the numeric rectangle.
            header_row: int | None = None
            for up in range(1, 7):
                rr = min_row - up
                if rr < 1:
                    break
                header_vals = [ws.cell(row=rr, column=min_col + k).value for k in range(width)]
                text_n = sum(1 for v in header_vals if isinstance(v, str) and v.strip())
                num_n = sum(1 for v in header_vals if _is_number(v))
                if text_n >= 2 and num_n == 0:
                    header_row = rr
                    break

            start_row = header_row if header_row is not None else min_row

            # Expand left by up to 2 cols when:
            # - header row has text there, AND
            # - the column has enough numeric coverage in the detected block.
            start_col = min_col
            if header_row is not None:
                block_h = max(1, max_row - min_row + 1)
                for left in range(1, 3):
                    cc = min_col - left
                    if cc < 1:
                        break
                    hv = ws.cell(row=header_row, column=cc).value
                    if not (isinstance(hv, str) and hv.strip()) or _is_number(hv):
                        continue
                    cov = sum(1 for rrr in range(min_row, max_row + 1) if _is_number(ws.cell(row=rrr, column=cc).value))
                    if cov / float(block_h) < 0.5:
                        continue
                    start_col = cc
            end_row = max_row
            end_col = max_col

            a1 = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            extracted = extract_a1_range(ws, a1, max_rows=80, max_cols=20)
            total_cells = max(1, (end_row - start_row + 1) * (end_col - start_col + 1))
            numeric_cells = sum(
                1
                for rr in range(start_row, end_row + 1)
                for cc in range(start_col, end_col + 1)
                if _is_number(ws.cell(row=rr, column=cc).value)
            )
            candidates.append(
                ExcelBlockCandidate(
                    sheet=ws.title,
                    a1_range=a1,
                    numeric_cells=numeric_cells,
                    total_cells=total_cells,
                    preview_rows=extracted[:6],
                )
            )

    # Prefer larger / denser numeric blocks.
    candidates_sorted = sorted(
        candidates,
        key=lambda c: (c.numeric_cells, c.numeric_cells / float(max(1, c.total_cells))),
        reverse=True,
    )
    return candidates_sorted[:max_blocks]


def find_numeric_blocks(
    ws,
    *,
    max_scan_rows: int = 220,
    max_scan_cols: int = 32,
    max_blocks: int = 10,
    extra_windows: int = 3,
) -> list[ExcelBlockCandidate]:
    """
    Numeric block detection that works even when the table is far down the sheet:
    - Always scan top-left window (fast, common case).
    - Additionally, find dense numeric regions via sparse numeric cell coordinates and scan those windows.
    """
    # 1) Top-left scan.
    all_candidates: list[ExcelBlockCandidate] = _find_numeric_blocks_in_window(
        ws,
        start_row=1,
        start_col=1,
        height=max_scan_rows,
        width=max_scan_cols,
        max_blocks=max_blocks,
    )

    # 2) Extra windows based on numeric density.
    points = _numeric_cells(ws)
    if not points:
        return _dedupe_candidates(all_candidates)[:max_blocks]

    win_h = max_scan_rows
    win_w = max_scan_cols
    buckets: dict[tuple[int, int], int] = {}
    for r, c in points:
        br = (max(1, r) - 1) // win_h
        bc = (max(1, c) - 1) // win_w
        buckets[(br, bc)] = buckets.get((br, bc), 0) + 1

    top = sorted(buckets.items(), key=lambda kv: kv[1], reverse=True)
    picked = 0
    for (br, bc), _count in top:
        if br == 0 and bc == 0:
            continue
        start_r = br * win_h + 1
        start_c = bc * win_w + 1
        all_candidates.extend(
            _find_numeric_blocks_in_window(
                ws,
                start_row=start_r,
                start_col=start_c,
                height=win_h,
                width=win_w,
                max_blocks=max_blocks,
            )
        )
        picked += 1
        if picked >= extra_windows:
            break

    all_candidates = _dedupe_candidates(all_candidates)
    all_candidates_sorted = sorted(
        all_candidates,
        key=lambda c: (c.numeric_cells, c.numeric_cells / float(max(1, c.total_cells))),
        reverse=True,
    )
    return all_candidates_sorted[:max_blocks]


@dataclass(frozen=True)
class ExcelChartSeriesRef:
    sheet: str
    values_range: str
    x_range: str | None = None
    title: str | None = None


@dataclass(frozen=True)
class ExcelChartRef:
    chart_id: str
    sheet: str
    chart_type: str
    title: str | None
    series: list[ExcelChartSeriesRef]


def _parse_ref(ref: str) -> tuple[str, str] | None:
    """
    Parse openpyxl chart references like:
      "'Sheet 1'!$A$1:$A$10" or "Sheet1!$B$2:$B$20"
    Returns (sheet_name, A1_RANGE).
    """
    s = (ref or "").strip()
    if "!" not in s:
        return None
    sheet_raw, a1 = s.split("!", 1)
    sheet = sheet_raw.strip().strip("'").strip('"')
    a1 = a1.replace("$", "").strip()
    if not validate_a1_range(a1):
        return None
    return sheet, a1


def list_chart_refs(wb, *, max_charts: int = 20, max_series: int = 20) -> list[ExcelChartRef]:
    """
    Extract chart -> series -> cell-range references (no rendering).
    This enables re-plotting with matplotlib deterministically.
    """
    charts: list[ExcelChartRef] = []
    for ws in wb.worksheets:
        chart_objs = getattr(ws, "_charts", None) or []
        if not chart_objs:
            continue
        for chart_idx, chart in enumerate(chart_objs[:max_charts]):
            chart_type = type(chart).__name__
            title = None
            try:
                if chart.title and getattr(chart.title, "tx", None):
                    # Best-effort to get the first run text.
                    title = chart.title.tx.rich.p[0].r[0].t  # type: ignore[attr-defined]
            except Exception:
                title = None

            series_refs: list[ExcelChartSeriesRef] = []
            for s_idx, ser in enumerate(getattr(chart, "series", [])[:max_series]):
                values_ref = None
                x_ref = None
                try:
                    values_ref = getattr(getattr(getattr(ser, "val", None), "numRef", None), "f", None)
                except Exception:
                    values_ref = None
                try:
                    x_ref = getattr(getattr(getattr(ser, "xVal", None), "numRef", None), "f", None)
                except Exception:
                    x_ref = None
                if not values_ref:
                    try:
                        values_ref = getattr(getattr(getattr(ser, "val", None), "strRef", None), "f", None)
                    except Exception:
                        values_ref = None
                if not x_ref:
                    try:
                        x_ref = getattr(getattr(getattr(ser, "cat", None), "strRef", None), "f", None)
                    except Exception:
                        x_ref = None

                parsed_val = _parse_ref(str(values_ref)) if values_ref else None
                if not parsed_val:
                    continue
                val_sheet, val_range = parsed_val

                parsed_x = _parse_ref(str(x_ref)) if x_ref else None
                x_range = parsed_x[1] if parsed_x and parsed_x[0] == val_sheet else None

                ser_title = None
                try:
                    if getattr(ser, "title", None) and getattr(ser.title, "tx", None):
                        ser_title = ser.title.tx.rich.p[0].r[0].t  # type: ignore[attr-defined]
                except Exception:
                    ser_title = None

                series_refs.append(
                    ExcelChartSeriesRef(sheet=val_sheet, values_range=val_range, x_range=x_range, title=ser_title)
                )

            if not series_refs:
                continue
            chart_id = f"{ws.title}#{chart_idx}"
            charts.append(
                ExcelChartRef(
                    chart_id=chart_id,
                    sheet=ws.title,
                    chart_type=chart_type,
                    title=title,
                    series=series_refs,
                )
            )
    return charts


def build_table_from_chart(wb, chart: ExcelChartRef, *, max_rows: int = 200) -> list[list[str]]:
    """
    Create a 2D table: [x, y1, y2, ...] from chart refs.
    If x_range is missing, row index is used as x.
    """
    if not chart.series:
        return []
    ws = wb[chart.series[0].sheet]

    # Extract x.
    x_vals: list[str] = []
    x_range = chart.series[0].x_range
    if x_range:
        x_col = extract_a1_range(ws, x_range, max_rows=max_rows, max_cols=1)
        x_vals = [r[0] if r else "" for r in x_col]
    else:
        # Use 1..N based on first series length.
        y0 = extract_a1_range(ws, chart.series[0].values_range, max_rows=max_rows, max_cols=1)
        x_vals = [str(i + 1) for i in range(len(y0))]

    # Extract y series.
    y_series: list[tuple[str, list[str]]] = []
    for ser in chart.series:
        ws2 = wb[ser.sheet]
        col = extract_a1_range(ws2, ser.values_range, max_rows=max_rows, max_cols=1)
        y = [r[0] if r else "" for r in col]
        name = (ser.title or "").strip() or f"series_{len(y_series) + 1}"
        y_series.append((name, y))

    n = min([len(x_vals)] + [len(y) for _, y in y_series]) if y_series else len(x_vals)
    if n <= 0:
        return []
    header = ["x"] + [name for name, _ in y_series]
    rows: list[list[str]] = [header]
    for i in range(n):
        row = [x_vals[i]] + [y[i] for _, y in y_series]
        rows.append(row)
    return _trim_empty(rows)
