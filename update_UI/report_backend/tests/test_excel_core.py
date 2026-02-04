from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

from core.excel import build_table_from_chart, find_numeric_blocks, list_chart_refs, load_workbook_bytes
from graph.nodes.excel_mvp import _infer_xy_series


def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_chart_refs_and_table_build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["V", "I"])
    for i in range(1, 6):
        ws.append([i, i * 2])

    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=6)  # includes header "I"
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)  # V values
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")

    wb2 = load_workbook_bytes(_to_bytes(wb))
    charts = list_chart_refs(wb2)
    assert charts, "Expected at least one chart reference"

    table = build_table_from_chart(wb2, charts[0])
    assert len(table) >= 3
    assert len(table[0]) >= 2
    # Header is best-effort; ensure body looks numeric-ish.
    assert table[1][0] != ""
    assert table[1][1] != ""


def test_find_numeric_blocks_includes_header_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["t", "v", "note"])
    for i in range(1, 7):
        ws.append([i, i * 10, ""])

    wb2 = load_workbook_bytes(_to_bytes(wb))
    ws2 = wb2["Sheet1"]
    blocks = find_numeric_blocks(ws2, max_blocks=5)
    assert blocks
    # Expect header row included, and 2 numeric cols at least.
    best = blocks[0]
    assert best.sheet == "Sheet1"
    assert "A1" in best.a1_range or best.a1_range.startswith("A1")


def test_find_numeric_blocks_finds_table_far_down_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Put the table far down so top-left scan alone would miss it.
    start_row = 520
    ws.cell(row=start_row, column=3, value="t")
    ws.cell(row=start_row, column=4, value="v")
    for i in range(1, 9):
        ws.cell(row=start_row + i, column=3, value=i)
        ws.cell(row=start_row + i, column=4, value=i * 5)

    wb2 = load_workbook_bytes(_to_bytes(wb))
    ws2 = wb2["Sheet1"]
    blocks = find_numeric_blocks(ws2, max_blocks=5, extra_windows=3)
    assert blocks
    # Expect a range that starts at row=start_row (header included).
    assert str(start_row) in blocks[0].a1_range


def test_find_numeric_blocks_finds_header_several_rows_above_dense_block():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row is intentionally separated from dense numeric block.
    ws.cell(row=6, column=2, value="Ib[mA]")
    ws.cell(row=6, column=3, value="Vbe[V]")
    ws.cell(row=6, column=4, value="Ib[mA]")
    ws.cell(row=6, column=5, value="Vbe[V]")

    # Dense numeric block starts at row=10 and is shifted right; row coverage is high there.
    for i in range(10, 18):
        ws.cell(row=i, column=3, value=0.40 + (i - 10) * 0.02)  # Vbe
        ws.cell(row=i, column=4, value=0.01 + (i - 10) * 0.05)  # Ib
        ws.cell(row=i, column=5, value=0.41 + (i - 10) * 0.02)  # Vbe
        ws.cell(row=i, column=6, value=0.02 + (i - 10) * 0.05)  # Ib

    wb2 = load_workbook_bytes(_to_bytes(wb))
    ws2 = wb2["Sheet1"]
    blocks = find_numeric_blocks(ws2, max_blocks=5)
    assert blocks
    # Expect the detected block to include the header row at row=6.
    assert "6" in blocks[0].a1_range


def test_infer_xy_series_swaps_current_voltage_pair():
    # Header is (Ib[mA], Vbe[V]) pairs; we want x=Vbe[V], y=Ib[mA] in the output.
    table_rows = [
        ["Ib[mA]", "Vbe[V]", "Ib[mA]", "Vbe[V]"],
        ["0.01", "0.40", "0.02", "0.41"],
        ["0.06", "0.50", "0.07", "0.51"],
    ]
    meta, series = _infer_xy_series(table_rows, driver=None)
    assert meta["x_name"].lower() == "vbe"
    assert meta["x_unit"].lower() == "v"
    assert meta["y_name"].lower() == "ib"
    assert meta["y_unit"].lower() == "ma"
    assert series and series[0]["points"][0][0] == 0.40 and series[0]["points"][0][1] == 0.01
